"""Build Excel output and markdown summary."""
import pandas as pd
from build_system import main, LEVELS, POSITIONS, ROSTER_SIZES, is_catcher, projected_pos_adj, is_high_potential, woba_max_level
from build_pitcher_system import main as pitcher_main, SP_PER_LEVEL, RP_PER_LEVEL, PITCHER_ROSTER_SIZE, pwoba_top_level
from org_report import build_batting_order, estimate_runs_per_game
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def _hitter_top_level(p):
    """Highest level a hitter could play given current wOBA (with the
    PREMIUM_WOBA_RELAX rule applied via woba_max_level). Returned as the
    level abbreviation. Used on the Release Pool to show what slot would
    fit this player if one were available."""
    try:
        return LEVELS[woba_max_level(p)]
    except Exception:
        return '?'


def _pitcher_top_level(p):
    """Highest level for a pitcher = current pwOBA ceiling (mirrors the
    _top calculation used in build_pitcher_system, which is now pwOBA-only
    after the age-cap removal). Pitchers with no pwOBA data default to
    R(DLR)."""
    try:
        return LEVELS[min(pwoba_top_level(p), len(LEVELS) - 1)]
    except Exception:
        return '?'


def _platoon_lineup_extras(split_starters, vs_key):
    """Compute (name -> batting-order slot) and runs/game estimate for a
    9-position starting nine dict, using the same logic as the org-report
    HTML used to. Skips empty positions. Returns ({}, NaN) if the lineup
    is empty."""
    rows = []
    for pos, p in split_starters.items():
        if p is None:
            continue
        rows.append({
            'name': p['name'],
            'pos': pos,
            'wOBA_vs': p.get(vs_key) or 0,
        })
    if not rows:
        return {}, float('nan')
    df = pd.DataFrame(rows)
    side = 'R' if vs_key == 'wOBAR' else 'L'
    order_df = build_batting_order(df, side=side)
    name_to_slot = dict(zip(order_df['name'], order_df['slot']))
    rpg = estimate_runs_per_game(order_df)
    return name_to_slot, rpg

OUTFILE_TEMPLATE = 'outputs/{org}_roster_system.xlsx'

# Styles
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill('solid', start_color='1F4E78')
SECTION_FONT = Font(name='Arial', bold=True, size=11, color='FFFFFF')
SECTION_FILL = PatternFill('solid', start_color='2E75B6')
DEFAULT_FONT = Font(name='Arial', size=10)
PROSPECT_FILL = PatternFill('solid', start_color='E2EFDA')  # light green for high-pot prospects
THIN = Side(border_style='thin', color='CCCCCC')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

LEVEL_COLORS = {
    'MLB': 'C00000',
    'AAA': 'ED7D31',
    'AA': 'FFC000',
    'A+': '70AD47',
    'A': '5B9BD5',
    'R': '7030A0',
    'R(DLR)': '595959',
}

def write_level_sheet(ws, lvl, roster):
    starters = roster['starters']
    bench = roster['bench']
    
    ws['A1'] = f'{lvl} Roster'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', start_color=LEVEL_COLORS[lvl])
    ws.merge_cells('A1:I1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24
    
    use_projected = lvl not in ('MLB', 'AAA')
    pos_label = 'projected pos_adj' if use_projected else 'pos_adj'
    headers = ['Slot', 'Name', 'Age', 'Pos', 'wOBA', 'wOBAP', 'Gap', pos_label, 'Notes']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal='center')
        c.border = BORDER
    
    row = 4
    sec = ws.cell(row=row, column=1, value='STARTERS')
    sec.font = SECTION_FONT
    sec.fill = SECTION_FILL
    sec.alignment = Alignment(horizontal='center')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    row += 1
    
    def write_player_row(row, p, slot_label):
        if p:
            woba = p.get('wOBA') or 0
            wobap = p.get('wOBAP') or 0
            gap = wobap - woba
            # Prospect labelling tracks is_high_potential() (minor=1, age ≤ HP_MAX_AGE,
            # wOBAP above the position-specific HP threshold) rather than raw gap.
            # A big wOBA→wOBAP gap on a player whose wOBAP is still below the HP
            # floor doesn't make them a prospect — just a weak bat with some upside.
            note = ''
            is_hp = is_high_potential(p)
            if is_hp: note = 'High-potential prospect'
            elif woba >= 0.320: note = 'Above-avg MLB bat'
            ws.cell(row=row, column=1, value=slot_label)
            ws.cell(row=row, column=2, value=p['name'])
            ws.cell(row=row, column=3, value=p['age'])
            ws.cell(row=row, column=4, value=p['pos_adj'])
            ws.cell(row=row, column=5, value=round(woba, 3))
            ws.cell(row=row, column=6, value=round(wobap, 3))
            ws.cell(row=row, column=7, value=round(gap, 3))
            # Pos value: at minor levels show projected (current pos_adj + bat dev), else current
            if use_projected:
                # Use projected at player's primary position
                pos_val = projected_pos_adj(p, p['pos_adj']) or 0
            else:
                pos_val = p.get('best_adj') or 0
            ws.cell(row=row, column=8, value=round(pos_val, 2))
            ws.cell(row=row, column=9, value=note)
            if is_hp:
                for col in range(1, 10):
                    ws.cell(row=row, column=col).fill = PROSPECT_FILL
            elif woba >= 0.320:
                for col in range(1, 10):
                    ws.cell(row=row, column=col).fill = PatternFill('solid', start_color='FFF2CC')
        else:
            ws.cell(row=row, column=1, value=slot_label)
            ws.cell(row=row, column=2, value='-- empty --').font = Font(name='Arial', italic=True, color='999999')
        ws.cell(row=row, column=1).font = Font(name='Arial', bold=True, size=10)
        for col in range(1, 10):
            ws.cell(row=row, column=col).border = BORDER
            if not ws.cell(row=row, column=col).font.bold:
                ws.cell(row=row, column=col).font = DEFAULT_FONT
    
    for pos in POSITIONS:
        write_player_row(row, starters.get(pos), pos)
        row += 1

    # vs RHP and vs LHP lineup variants on the same roster.
    # Compact rendering: position / name / age / vs-handedness wOBA / overall
    # wOBA / platoon delta / [Δ vs standard]. The last column flags positions
    # where the platoon optimum differs from the standard starting nine.
    def write_platoon_block(row, label, split_starters, vs_key):
        # Compute batting order (1-9 by The-Book heuristic) and an estimated
        # R/G for this lineup. Migrated from the old org_report.html so the
        # xlsx is now the single source of MLB lineup detail.
        name_to_slot, rpg = _platoon_lineup_extras(split_starters, vs_key)

        sec = ws.cell(row=row, column=1, value=label)
        sec.font = SECTION_FONT
        sec.fill = SECTION_FILL
        sec.alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        row += 1
        # mini-header — Order column added; rows are still position-ordered
        # for visual continuity with the standard starters block.
        headers = ['Order', 'Pos', 'Name', 'Age', vs_key, 'wOBA', 'Δ', '', 'Note']
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = Font(name='Arial', bold=True, size=9, color='666666')
            c.alignment = Alignment(horizontal='center')
            c.border = BORDER
        row += 1
        for pos in POSITIONS:
            p = split_starters.get(pos)
            std = starters.get(pos)
            slot = name_to_slot.get(p['name']) if p else None
            ws.cell(row=row, column=1, value=slot if slot is not None else '').font = Font(name='Arial', bold=True, size=10)
            ws.cell(row=row, column=2, value=pos).font = Font(name='Arial', bold=True, size=10)
            if p is None:
                ws.cell(row=row, column=3, value='-- empty --').font = Font(name='Arial', italic=True, color='999999')
            else:
                woba = p.get('wOBA') or 0
                split_woba = p.get(vs_key) or 0
                delta = split_woba - woba
                ws.cell(row=row, column=3, value=p['name'])
                ws.cell(row=row, column=4, value=p['age'])
                ws.cell(row=row, column=5, value=round(split_woba, 3))
                ws.cell(row=row, column=6, value=round(woba, 3))
                ws.cell(row=row, column=7, value=round(delta, 3))
                if std is None or std['name'] != p['name']:
                    ws.cell(row=row, column=9, value=f'swap from {std["name"] if std else "(empty)"}').font = Font(name='Arial', italic=True, size=9, color='C00000')
            for col in range(1, 10):
                ws.cell(row=row, column=col).border = BORDER
                if not ws.cell(row=row, column=col).font.bold:
                    ws.cell(row=row, column=col).font = DEFAULT_FONT
            row += 1
        # R/G footer — italic, spans the block.
        if not (isinstance(rpg, float) and (rpg != rpg)):  # not NaN
            ws.cell(row=row, column=1, value=f'Est. R/G: {rpg:.2f}').font = Font(name='Arial', italic=True, size=9, color='666666')
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
            row += 1
        return row

    row = write_platoon_block(row, 'VS RHP', roster.get('starters_vsR', {}), 'wOBAR')
    row = write_platoon_block(row, 'VS LHP', roster.get('starters_vsL', {}), 'wOBAL')

    sec = ws.cell(row=row, column=1, value='BENCH / DEPTH')
    sec.font = SECTION_FONT
    sec.fill = SECTION_FILL
    sec.alignment = Alignment(horizontal='center')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    row += 1

    # Bench is now ordered by named role: Backup C, Utility IF, Utility OF,
    # Best bat, then depth. Use the slot label as both the leftmost column
    # and the Notes column for clarity.
    for role, p in roster['bench_roles']:
        write_player_row(row, p, role)
        if p is None:
            ws.cell(row=row, column=9, value=f'(no {role.lower()} candidate)')
        else:
            if role == 'Depth' and is_high_potential(p):
                ws.cell(row=row, column=9, value='Prospect (depth)')
            else:
                ws.cell(row=row, column=9, value=role)
        row += 1
    
    widths = [8, 26, 6, 7, 8, 9, 7, 8, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = w
    
    row += 1
    ws.cell(row=row, column=1, value=f'Total: {len(roster["all"])} players (target {ROSTER_SIZES[lvl]})').font = Font(name='Arial', italic=True, size=9, color='666666')

def write_summary(ws, rosters, overflow, org='LAA'):
    ws['A1'] = f'{org} Hitter System - Summary'
    ws['A1'].font = Font(name='Arial', bold=True, size=14)
    ws.merge_cells('A1:F1')
    
    headers = ['Level', 'Total', 'Catchers', 'Top Prospect', 'Avg Best', 'Avg BestP']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal='center')
        c.border = BORDER
    
    row = 4
    for lvl in LEVELS:
        all_p = rosters[lvl]['all']
        cs = [p for p in all_p if is_catcher(p)]
        top = max(all_p, key=lambda p: p['bestP'] or -99)
        avg_best = sum(p['best'] or 0 for p in all_p) / len(all_p) if all_p else 0
        avg_pot = sum(p['bestP'] or 0 for p in all_p) / len(all_p) if all_p else 0
        ws.cell(row=row, column=1, value=lvl).font = Font(name='Arial', bold=True, color=LEVEL_COLORS[lvl])
        ws.cell(row=row, column=2, value=len(all_p))
        ws.cell(row=row, column=3, value=len(cs))
        ws.cell(row=row, column=4, value=f"{top['name']} ({top['age']}, {top['bestP']:.1f})")
        ws.cell(row=row, column=5, value=round(avg_best, 2))
        ws.cell(row=row, column=6, value=round(avg_pot, 2))
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = BORDER
            if col != 1: ws.cell(row=row, column=col).font = DEFAULT_FONT
        row += 1
    
    # Overflow row
    ws.cell(row=row, column=1, value='Release Pool').font = Font(name='Arial', bold=True, color='999999')
    ws.cell(row=row, column=2, value=len(overflow))
    ws.cell(row=row, column=3, value=sum(1 for p in overflow if is_catcher(p)))
    if overflow:
        top = max(overflow, key=lambda p: p['bestP'] or -99)
        ws.cell(row=row, column=4, value=f"{top['name']} ({top['age']}, {top['bestP']:.1f})")
    avg_best = sum(p['best'] or 0 for p in overflow) / len(overflow) if overflow else 0
    avg_pot = sum(p['bestP'] or 0 for p in overflow) / len(overflow) if overflow else 0
    ws.cell(row=row, column=5, value=round(avg_best, 2))
    ws.cell(row=row, column=6, value=round(avg_pot, 2))
    for col in range(1, 7):
        ws.cell(row=row, column=col).border = BORDER
        if col != 1: ws.cell(row=row, column=col).font = DEFAULT_FONT
    
    widths = [14, 8, 10, 32, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = w
    
    # Methodology notes
    row += 3
    ws.cell(row=row, column=1, value='Methodology').font = Font(name='Arial', bold=True, size=12)
    row += 1
    notes = [
        'Hybrid level assignment combines:',
        '  - Age developmental floor (typical age-by-level expectations)',
        '  - Current ability ceiling (best WAR projection prevents overmatching)',
        '  - Potential bump (top prospects pushed up 1 level)',
        '',
        'Roster sizes: MLB/AAA/AA/A+/A = 13, R/R(DLR) = 15. Hitters only (95 total).',
        '',
        'Catchers: 2 per level (starter + backup), pre-allocated before non-catchers.',
        '',
        'Starter selection:',
        '  - MLB & AAA: position-adjusted WAR drives selection (performance focus)',
        '  - AA and below: bestP drives selection (development focus)',
        '',
        'Light-green rows = top prospects (bestP >= 3.0).',
    ]
    for n in notes:
        ws.cell(row=row, column=1, value=n).font = Font(name='Arial', size=10)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1

def write_overflow(ws, overflow):
    ws['A1'] = 'Release / Overflow Pool'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', start_color='595959')
    ws.merge_cells('A1:F1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

    headers = ['Name', 'Age', 'Pos', 'Top level', 'Best', 'BestP']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER

    row = 4
    overflow_sorted = sorted(overflow, key=lambda p: (p['bestP'] or -99), reverse=True)
    for p in overflow_sorted:
        ws.cell(row=row, column=1, value=p['name'])
        ws.cell(row=row, column=2, value=p['age'])
        ws.cell(row=row, column=3, value=p['pos_adj'])
        ws.cell(row=row, column=4, value=_hitter_top_level(p))
        ws.cell(row=row, column=5, value=round(p['best'], 1))
        ws.cell(row=row, column=6, value=round(p['bestP'], 1))
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = BORDER
            ws.cell(row=row, column=col).font = DEFAULT_FONT
        row += 1

    widths = [26, 6, 8, 10, 10, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = w

def write_pitcher_sheet(ws, lvl, roster):
    """One sheet per level for pitchers: 5-man rotation, 8-man bullpen."""
    ws['A1'] = f'{lvl} Pitching Staff'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', start_color=LEVEL_COLORS[lvl])
    ws.merge_cells('A1:H1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24

    headers = ['Slot', 'Name', 'Age', 'pwOBA', 'pwOBAP', 'Role WAR', 'IP', 'Notes']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal='center')
        c.border = BORDER

    def sp_note(spp):
        if spp >= 2.0: return 'Front-of-rotation'
        if spp >= 0.5: return 'Mid-rotation'
        if spp >= 0.0: return 'Back-end starter'
        return 'Innings eater'

    def rp_note(rpp):
        if rpp >= 0.8: return 'High leverage'
        if rpp >= 0.0: return 'Mid leverage'
        return 'Low leverage / depth'

    def write_pitcher_row(row, p, slot_label, war_key, note_fn):
        ws.cell(row=row, column=1, value=slot_label).font = Font(name='Arial', bold=True, size=10)
        ws.cell(row=row, column=2, value=p['name'])
        ws.cell(row=row, column=3, value=p['age'])
        ws.cell(row=row, column=4, value=round(p.get('pwOBA') or 0, 3))
        ws.cell(row=row, column=5, value=round(p.get('pwOBAP') or 0, 3))
        war = p.get(war_key)
        ws.cell(row=row, column=6, value=round(war, 2) if war is not None else '')
        ws.cell(row=row, column=7, value=p.get('ip', 0))
        ws.cell(row=row, column=8, value=note_fn(war if war is not None else -99))
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = BORDER
            if not ws.cell(row=row, column=col).font.bold:
                ws.cell(row=row, column=col).font = DEFAULT_FONT

    row = 4
    sec = ws.cell(row=row, column=1, value='STARTING ROTATION')
    sec.font = SECTION_FONT
    sec.fill = SECTION_FILL
    sec.alignment = Alignment(horizontal='center')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 1

    for i in range(SP_PER_LEVEL):
        if i < len(roster['starters']):
            write_pitcher_row(row, roster['starters'][i], f'SP{i+1}', 'sp_warP', sp_note)
        else:
            ws.cell(row=row, column=1, value=f'SP{i+1}').font = Font(name='Arial', bold=True, size=10)
            ws.cell(row=row, column=2, value='-- empty --').font = Font(name='Arial', italic=True, color='999999')
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = BORDER
        row += 1

    sec = ws.cell(row=row, column=1, value='BULLPEN')
    sec.font = SECTION_FONT
    sec.fill = SECTION_FILL
    sec.alignment = Alignment(horizontal='center')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 1

    for i in range(RP_PER_LEVEL):
        if i < len(roster['bullpen']):
            write_pitcher_row(row, roster['bullpen'][i], f'RP{i+1}', 'rp_warP', rp_note)
        else:
            ws.cell(row=row, column=1, value=f'RP{i+1}').font = Font(name='Arial', bold=True, size=10)
            ws.cell(row=row, column=2, value='-- empty --').font = Font(name='Arial', italic=True, color='999999')
            ws.cell(row=row, column=8, value='Sign FA').font = Font(name='Arial', italic=True, size=9, color='C00000')
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = BORDER
        row += 1

    widths = [8, 26, 6, 9, 10, 11, 7, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = w

    row += 1
    ws.cell(row=row, column=1, value=f'Total: {len(roster["all"])} pitchers (target {PITCHER_ROSTER_SIZE})').font = Font(name='Arial', italic=True, size=9, color='666666')


def write_pitcher_overflow(ws, overflow):
    ws['A1'] = 'Pitcher Release / Overflow Pool'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', start_color='595959')
    ws.merge_cells('A1:G1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

    headers = ['Name', 'Age', 'Top level', 'pwOBA', 'sp_warP', 'rp_warP', 'IP']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER

    row = 4
    overflow_sorted = sorted(overflow, key=lambda p: -(p.get('rp_warP') or -99))
    for p in overflow_sorted:
        ws.cell(row=row, column=1, value=p['name'])
        ws.cell(row=row, column=2, value=p['age'])
        ws.cell(row=row, column=3, value=_pitcher_top_level(p))
        ws.cell(row=row, column=4, value=round(p.get('pwOBA') or 0, 3))
        sp = p.get('sp_warP')
        ws.cell(row=row, column=5, value=round(sp, 2) if sp is not None else '')
        rp = p.get('rp_warP')
        ws.cell(row=row, column=6, value=round(rp, 2) if rp is not None else '')
        ws.cell(row=row, column=7, value=p.get('ip', 0))
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = BORDER
            ws.cell(row=row, column=col).font = DEFAULT_FONT
        row += 1

    widths = [26, 6, 10, 9, 10, 10, 7]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = w


def write_flagged(ws, players, kind):
    """Render players pulled out of active placement via flagged.txt — typically
    injuries. Re-run the system once they're cleared from flagged.txt."""
    ws['A1'] = f'Flagged / Unavailable {kind}'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', start_color='8B0000')
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

    ws['A3'] = 'Name'
    ws['B3'] = 'Age'
    ws['C3'] = 'Pos'
    ws['D3'] = 'Note'
    for col in range(1, 5):
        ws.cell(row=3, column=col).font = HEADER_FONT
        ws.cell(row=3, column=col).fill = HEADER_FILL
        ws.cell(row=3, column=col).border = BORDER

    row = 4
    for p in sorted(players, key=lambda x: x['name']):
        ws.cell(row=row, column=1, value=p['name'])
        ws.cell(row=row, column=2, value=p['age'])
        ws.cell(row=row, column=3, value=p.get('pos_adj') or p.get('pos') or '')
        ws.cell(row=row, column=4, value='Flagged in flagged.txt — rerun once cleared')
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = BORDER
            ws.cell(row=row, column=col).font = DEFAULT_FONT
        row += 1

    widths = [26, 6, 6, 48]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = w


def main_build(org=None):
    if org is None:
        from config import team_managed
        org = team_managed
    rosters, overflow, flagged = main(org=org)
    pitcher_rosters, pitcher_overflow, pitcher_flagged = pitcher_main(org=org)
    wb = Workbook()
    wb.remove(wb.active)

    # Summary first
    ws = wb.create_sheet('Summary')
    write_summary(ws, rosters, overflow, org=org)

    # Hitter level sheets
    for lvl in LEVELS:
        sheet_name = lvl.replace('(', '_').replace(')', '')
        ws = wb.create_sheet(sheet_name)
        write_level_sheet(ws, lvl, rosters[lvl])

    # Pitcher level sheets — same level order, suffix _P to keep names short
    # (sheet names are capped at 31 chars by Excel; with _P we stay well under).
    for lvl in LEVELS:
        sheet_name = lvl.replace('(', '_').replace(')', '') + '_P'
        ws = wb.create_sheet(sheet_name)
        write_pitcher_sheet(ws, lvl, pitcher_rosters[lvl])

    # Overflow + flagged pools
    ws = wb.create_sheet('Release_Pool')
    write_overflow(ws, overflow)
    ws = wb.create_sheet('Release_Pool_P')
    write_pitcher_overflow(ws, pitcher_overflow)
    if flagged:
        ws = wb.create_sheet('Flagged')
        write_flagged(ws, flagged, 'Hitters')
    if pitcher_flagged:
        ws = wb.create_sheet('Flagged_P')
        write_flagged(ws, pitcher_flagged, 'Pitchers')

    outfile = OUTFILE_TEMPLATE.format(org=org)
    wb.save(outfile)
    print(f'Saved: {outfile}')

if __name__ == '__main__':
    main_build()
